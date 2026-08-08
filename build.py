#!/usr/bin/env python3
"""
CBEDS site data builder.

Reads the master workbook (draft/CBEDSync.xlsx) and regenerates
cbedsync-data.js - the file the website reads to draw the network.

Run it after you edit the Excel:  python build.py
Or double-click  update-website.bat (Windows) / update-website.command (Mac).
"""

import json
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is not installed. Run:  pip install openpyxl")

HERE = Path(__file__).resolve().parent
XLSX = HERE / "draft" / "CBEDSync.xlsx"
OUT = HERE / "cbedsync-data.js"
DESC_MAX = 240
SUB_MAX = 60
WEB_MAX = 300

# "Source" says where a row came from: blank when the team added it, "Public" when it
# arrived through a form on the website and was approved. Unlike every other field
# here, it is found by its heading rather than its position, so it can sit in any
# column and can be moved later without touching this file.
SOURCE_HEAD = "Source"

# "Show/Hide" takes a row off the site without deleting it. Blank or "Show" publishes
# it as before; "Hide" leaves the row in the workbook but out of cbedsync-data.js.
# Found by heading too, so it can sit in any column and be moved later.
SHOW_HEADS = ("Show/Hide", "Show", "Visible")
HIDE_VALUES = {"hide", "hidden", "no", "n", "false", "0", "off"}
SHOW_VALUES = {"", "show", "shown", "visible", "yes", "y", "true", "1", "on"}

THEMES = [
    "Data Integration and Interoperability",
    "Economic and Market Transparency",
    "Compliance and Quality Assurance",
    "Lifecycle and Asset Performance",
    "Health, Safety, and Wellbeing",
    "Production and Construction Management",
    "Sustainability and Circularity",
]
STAGES = [
    "Data Needs and Requirements ",
    "Data Collection and Exchange",
    "Data Models and Integration",
    "Data Governance and Security",
    "Data Reporting and Analytics",
]


def s(v):
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        v = int(v)
    return str(v).strip()


def txt(v, maxlen=None):
    """Clean text: collapse newlines/tabs to spaces, optionally trim to maxlen + …"""
    d = s(v)
    for ch in ("\r\n", "\n", "\r", "\t"):
        d = d.replace(ch, " ")
    if maxlen is not None and len(d) > maxlen:
        d = d[:maxlen] + "…"
    return d


def find_col(ws, head):
    """Index of the column with this heading in row 1, or None if there is not one.

    Everything else on these sheets is read by position, which is why a column must
    never be inserted in the middle of one. This is looked up by name instead, so
    the Source column can live anywhere and be moved without breaking the build."""
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        for i, v in enumerate(row):
            if s(v).lower() == head.lower():
                return i
    return None


def find_col_any(ws, heads):
    """The first of these headings that the sheet actually has."""
    for head in heads:
        col = find_col(ws, head)
        if col is not None:
            return col
    return None


def is_hidden(row, col, sheet, odd):
    """True when the row is marked Hide. Blank means show, so a sheet nobody has
    touched publishes exactly as it did before this column existed.

    Anything that is neither is recorded and the row is shown. Hiding on a value the
    build did not understand would drop an entity silently, and a typo that shows one
    row too many is a far smaller problem than one that loses it."""
    if col is None or col >= len(row):
        return False
    v = s(row[col]).lower()
    if v in HIDE_VALUES:
        return True
    if v not in SHOW_VALUES:
        odd.append('%s "%s" = "%s"' % (sheet, s(row[0]), s(row[col])))
    return False


def source(row, col):
    """The Source cell for this row, or "" if the sheet has no such column.

    Rows come back only as wide as the sheet has data, so a trailing empty cell can
    make a row shorter than the heading - that is a blank Source, not an error."""
    if col is None or col >= len(row):
        return ""
    return s(row[col])


def flags(row, cols, labels):
    out = []
    for col, label in zip(cols, labels):
        if s(row[col]):
            out.append(label)
    return out


def values(row, cols):
    out = []
    for col in cols:
        v = s(row[col])
        if v and v not in out:
            out.append(v)
    return out


def rels(row, cols, t):
    out = []
    for col in cols:
        v = s(row[col])
        if v:
            out.append({"n": v, "t": t})
    return out


def build():
    if not XLSX.exists():
        sys.exit("Cannot find workbook: %s" % XLSX)
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    nodes = []
    no_source = []          # sheets with no Source column, reported at the end
    no_show = []            # sheets with no Show/Hide column
    odd_show = []           # cells that are neither Show nor Hide
    hidden_names = set()    # so links pointing at them can be reported
    hidden_count = 0

    ws = wb["Agent"]
    src_col = find_col(ws, SOURCE_HEAD)
    show_col = find_col_any(ws, SHOW_HEADS)
    if src_col is None:
        no_source.append("Agent")
    if show_col is None:
        no_show.append("Agent")
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = s(row[0])
        if not name:
            continue
        if is_hidden(row, show_col, "Agent", odd_show):
            hidden_names.add(name)
            hidden_count += 1
            continue
        node = {
            "id": name, "kind": "agent",
            "sub": txt(row[2], SUB_MAX), "cls": s(row[1]),
            "web": txt(row[3], WEB_MAX), "desc": txt(row[4], DESC_MAX), "loc": s(row[7]),
            "themes": flags(row, range(21, 28), THEMES),
            "stages": flags(row, range(28, 33), STAGES),
            "tech": values(row, range(33, 39)),
            "rel": rels(row, range(9, 21), "partOf") + rels(row, range(39, 55), "link"),
        }
        src = source(row, src_col)
        if src:                       # left out when blank, which is nearly every row
            node["src"] = src
        nodes.append(node)

    ws = wb["Project"]
    src_col = find_col(ws, SOURCE_HEAD)
    show_col = find_col_any(ws, SHOW_HEADS)
    if src_col is None:
        no_source.append("Project")
    if show_col is None:
        no_show.append("Project")
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = s(row[0])
        if not name:
            continue
        if is_hidden(row, show_col, "Project", odd_show):
            hidden_names.add(name)
            hidden_count += 1
            continue
        node = {
            "id": name, "kind": "project",
            "sub": s(row[2]), "cls": "Project",
            "web": txt(row[3], WEB_MAX), "desc": txt(row[1], DESC_MAX), "loc": s(row[6]),
            "themes": flags(row, range(9, 16), THEMES),
            "stages": flags(row, range(16, 21), STAGES),
            "tech": values(row, range(21, 27)),
            "rel": rels(row, [7, 8], "managedBy") + rels(row, range(27, 43), "link"),
        }
        src = source(row, src_col)
        if src:
            node["src"] = src
        nodes.append(node)

    ws = wb["Output"]
    src_col = find_col(ws, SOURCE_HEAD)
    show_col = find_col_any(ws, SHOW_HEADS)
    if src_col is None:
        no_source.append("Output")
    if show_col is None:
        no_show.append("Output")
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = s(row[0])
        if not name:
            continue
        if is_hidden(row, show_col, "Output", odd_show):
            hidden_names.add(name)
            hidden_count += 1
            continue
        node = {
            "id": name, "kind": "output",
            "sub": txt(row[2], SUB_MAX), "cls": s(row[1]),
            "web": txt(row[3], WEB_MAX), "desc": txt(row[5], DESC_MAX), "loc": "",
            "year": s(row[4]),
            "themes": flags(row, range(10, 17), THEMES),
            "stages": flags(row, range(17, 22), STAGES),
            "tech": values(row, range(22, 28)),
            "rel": rels(row, [8, 9], "producedBy") + rels(row, range(28, 44), "link"),
        }
        src = source(row, src_col)
        if src:
            node["src"] = src
        nodes.append(node)

    techcat = {}
    for row in wb["Technologies"].iter_rows(min_row=2, values_only=True):
        tech, cat = s(row[0]), s(row[1])
        if tech:
            techcat[tech] = cat

    counts = {
        "agent": sum(1 for n in nodes if n["kind"] == "agent"),
        "project": sum(1 for n in nodes if n["kind"] == "project"),
        "output": sum(1 for n in nodes if n["kind"] == "output"),
        "public": sum(1 for n in nodes if n.get("src") == "Public"),
    }
    # A name used by two entries is legitimate - a software company and the product
    # named after it - but anything that maps the graph by name has to expect it, and
    # for a long time the drawing code did not: one of a pair silently took the other's
    # place. Reported every run so a new one is noticed the day it appears.
    seen_names = {}
    for n in nodes:
        seen_names.setdefault(n["id"], []).append(n["kind"])
    shared = {k: v for k, v in seen_names.items() if len(v) > 1}

    # Hiding a row does not rewrite the rows that named it, so a LinksTo still pointing
    # at it now points at nothing. That is allowed - the graph simply does not draw it -
    # but it is worth saying out loud, because the usual reason to hide something is
    # that it should stop appearing, and this is the way it can keep appearing.
    dangling = sorted({r["n"] for n in nodes for r in n["rel"] if r["n"] in hidden_names})

    return {"nodes": nodes, "themes": THEMES, "stages": STAGES,
            "techcat": techcat, "counts": counts,
            "noSource": no_source, "sharedNames": shared,
            "noShow": no_show, "oddShow": odd_show,
            "hiddenCount": hidden_count, "danglingToHidden": dangling}


def main():
    data = build()
    missing = data.pop("noSource")          # a message for you, not data for the site
    shared = data.pop("sharedNames")
    no_show = data.pop("noShow")
    odd_show = data.pop("oddShow")
    hidden_count = data.pop("hiddenCount")
    dangling = data.pop("danglingToHidden")
    payload = json.dumps(data, ensure_ascii=False, separators=(",", ":"))
    OUT.write_text("window.CBEDS_DATA=" + payload, encoding="utf-8")
    c = data["counts"]
    print("OK  wrote %s" % OUT.name)
    print("    agents=%d  projects=%d  outputs=%d  total=%d"
          % (c["agent"], c["project"], c["output"], len(data["nodes"])))
    # A quick check that approved submissions were actually marked in the Source
    # column - if you approved one today and this still says 0, the cell is empty.
    print("    from public submissions=%d" % c["public"])
    # Hidden rows are still in the workbook, so this is the only place they are counted.
    # If it says 0 the morning after you hid something, the cell did not take.
    print("    hidden=%d  (kept in the Excel, left off the site)" % hidden_count)
    if dangling:
        print("    note: %d hidden %s still named by a link, which now draws nothing: %s"
              % (len(dangling), "entry is" if len(dangling) == 1 else "entries are",
                 "; ".join(dangling[:3]) + (" ..." if len(dangling) > 3 else "")))
    if odd_show:
        print('    note: %d %s neither Show nor Hide, so the row was published as '
              "usual: %s" % (len(odd_show),
                             'cell in "%s" says' % SHOW_HEADS[0] if len(odd_show) == 1
                             else 'cells in "%s" say' % SHOW_HEADS[0],
                             "; ".join(odd_show[:3])
                             + (" ..." if len(odd_show) > 3 else "")))
    if no_show:
        print('    note: no "%s" column on %s - add one at the far right of the sheet '
              "to hide an entry without deleting it" % (SHOW_HEADS[0], ", ".join(no_show)))
    if missing:
        print('    note: no "%s" column on %s - add one at the far right of the '
              "sheet to record which entries came from the public"
              % (SOURCE_HEAD, ", ".join(missing)))
    if shared:
        pairs = sorted("%s (%s)" % (k, "+".join(sorted(set(v))))
                       for k, v in shared.items())
        print("    names shared by more than one entry=%d" % len(shared))
        print("      %s" % "; ".join(pairs[:3])
              + (" ..." if len(pairs) > 3 else ""))
        print("      that is allowed - a company and its product often share a name -")
        print("      but relations naming one of these reach only the first of the pair.")


if __name__ == "__main__":
    main()
