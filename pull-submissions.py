#!/usr/bin/env python3
"""
CBEDS submission puller.

Fetches form submissions from Netlify and stages them in draft/Submissions.xlsx,
where they can be checked before anything is copied into the master workbook.

Run it:  python pull-submissions.py
Or double-click  get-submissions.bat

WHY A SEPARATE FILE
-------------------
draft/CBEDSync.xlsx holds formulas, three sets of cell comments with their legacy
drawings, and dozens of conditional-formatting rules. openpyxl cannot open and
re-save that without risking some of it. So this script only ever *reads* the
master - to copy its column layout - and only ever *writes* Submissions.xlsx, a
plain file it made itself, which is safe to rewrite every run.

HOW THE COLUMNS STAY IN STEP
----------------------------
The staging sheets are not a hardcoded copy of the master's columns. Each run
reads row 1 of Agent / Project / Output out of CBEDSync.xlsx and rebuilds the
staging headers from it, so the two can never drift. Columns A..<last> match the
master exactly: select that range on a staged row, copy, and paste it straight
into the master as a new row. One blank gutter column follows, then the review
columns, which are not part of the master and are never copied across.
"""

import json
import os
import re
import sys
import zipfile
import urllib.error
import urllib.request
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("openpyxl is not installed. Run:  pip install openpyxl")

HERE = Path(__file__).resolve().parent
MASTER = HERE / "draft" / "CBEDSync.xlsx"
STAGING = HERE / "draft" / "Submissions.xlsx"
API = "https://api.netlify.com/api/v1"

SHEETS = ("Agent", "Project", "Output")
SOURCE_HEAD = "Source"

# A "post to the CBEDSynergy network" is not a CBEDSync entity and has no sheet in
# the master, so it gets a small one of its own and is never copied across.
POST_COLS = ["Post", "Description", "Webpage"]

# Columns the reviewer works in. They sit past a blank gutter so that the master's
# own columns can be copied across in one block without dragging these along.
REVIEW = [
    "Review: Status",          # blank / approved / rejected - yours to fill in
    "Review: Notes",
    "Review: Submitted",       # when it arrived
    "Review: Submitted by",
    "Review: Email",
    "Review: Named lead",      # Charter only
    "Review: Commitments",     # Charter only
    "Review: Unplaced",        # anything with no column to go in, so none is lost
    "Review: Form",
    "Review: Netlify id",      # how a submission is recognised as already staged
]
ID_COL = "Review: Netlify id"

# Which submission field goes under which *heading* in the master. Headings rather
# than positions, so this survives the sheet being rearranged.
FIELD_TO_HEAD = {
    "desc": "Description",
    "link": "Webpage",
}
TECH_HEADS = ["Technology 1", "Technology2", "Technology 3",
              "Technology 4", "Technology 5", "Technology 6"]
LINK_HEADS = ["LinksTo %d" % i for i in range(1, 17)]
# A theme column is a flag: build.py counts it as set if the cell holds anything.
THEME_MARK = "x"


def env(name, default=None):
    """Read a variable from the environment or from .env next to this script."""
    if os.environ.get(name):
        return os.environ[name]
    envfile = HERE / ".env"
    if envfile.exists():
        for line in envfile.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            k, v = line.split("=", 1)
            if k.strip() == name:
                return v.strip().strip('"').strip("'")
    return default


def api(path, token):
    req = urllib.request.Request(API + path,
                                 headers={"Authorization": "Bearer " + token,
                                          "User-Agent": "cbeds-site"})
    try:
        with urllib.request.urlopen(req, timeout=30) as r:
            return json.loads(r.read().decode("utf-8"))
    except urllib.error.HTTPError as e:
        if e.code in (401, 403):
            sys.exit("Netlify refused the token (HTTP %d).\n"
                     "Check NETLIFY_TOKEN in .env - it needs to be a personal "
                     "access token from https://app.netlify.com/user/applications"
                     % e.code)
        if e.code == 404:
            sys.exit("Netlify has no such project (HTTP 404).\n"
                     "Check NETLIFY_SITE in .env. It should be the full domain, "
                     "e.g. cbeds.netlify.app")
        sys.exit("Netlify returned HTTP %d for %s" % (e.code, path))
    except urllib.error.URLError as e:
        sys.exit("Could not reach Netlify: %s" % e.reason)


def resolve_site(want, token):
    """Find the project, whether NETLIFY_SITE is its domain, its name or its API id.

    The API wants its own site id, but nobody has that to hand - what people know is
    the address the site answers on. So ask for the list and match against every name
    it might be known by, and if nothing matches say which projects the token can see
    rather than a bare 404."""
    sites = api("/sites", token)
    want = (want or "").strip().lower().rstrip("/")
    bare = want.replace("https://", "").replace("http://", "")
    for s in sites:
        known = {str(s.get(k) or "").lower().replace("https://", "").rstrip("/")
                 for k in ("id", "site_id", "name", "url", "ssl_url",
                           "default_domain", "custom_domain")}
        known.discard("")
        if bare in known or want in known or bare.split(".")[0] == str(s.get("name", "")).lower():
            return s["id"], (s.get("name") or bare)
    names = ", ".join(sorted(str(s.get("name")) for s in sites)) or "(none)"
    sys.exit('Could not find a Netlify project matching "%s".\n'
             "This token can see: %s\n"
             "Set NETLIFY_SITE in .env to one of those names." % (want, names))


def col_of(ref):
    """Column number from a cell reference: BM1 -> 65."""
    n = 0
    for ch in re.match(r"([A-Z]+)", ref).group(1):
        n = n * 26 + ord(ch) - 64
    return n


def header_row_from_xlsx(path, sheet_names):
    """Row 1 of each named sheet, read out of the file itself.

    openpyxl is not asked for this. A row with gaps - the master has twenty columns
    between LinksTo 16 and Source - comes back from read-only mode with its empty
    cells dropped and the survivors renumbered, and neither cell.column nor
    cell.coordinate survives that: Source arrived directly after the last heading
    instead of in BM, twice. The cell references in the XML are the one thing that
    cannot be renumbered, so they are what this reads."""
    out = {}
    with zipfile.ZipFile(path) as z:
        book = z.read("xl/workbook.xml").decode("utf-8", "replace")
        rels = z.read("xl/_rels/workbook.xml.rels").decode("utf-8", "replace")

        def attr(tag, name):
            m = re.search(r'\b%s="([^"]*)"' % name, tag)
            return m.group(1) if m else None

        # Each attribute is read on its own rather than in a fixed order. Excel does
        # not commit to one: the workbook this was written against put Id before
        # Target, the next one it saved put Target first, and a pattern expecting the
        # first order silently matched nothing and lost every sheet.
        target = {}
        for tag in re.findall(r"<Relationship\b[^>]*>", rels):
            rid, tgt = attr(tag, "Id"), attr(tag, "Target")
            if rid and tgt:
                target[rid] = tgt
        sheet_rid = {}
        for tag in re.findall(r"<sheet\b[^>]*>", book):
            nm, rid = attr(tag, "name"), attr(tag, "r:id")
            if nm and rid:
                sheet_rid[nm] = rid

        shared = []
        if "xl/sharedStrings.xml" in z.namelist():
            sx = z.read("xl/sharedStrings.xml").decode("utf-8", "replace")
            for si in re.findall(r"<si>(.*?)</si>", sx, re.S):
                shared.append("".join(re.findall(r"<t[^>]*>(.*?)</t>", si, re.S)))

        for name in sheet_names:
            rid = sheet_rid.get(name)
            if not rid or rid not in target:
                sys.exit('The master has no sheet called "%s".' % name)
            # targets come both ways too: "/xl/worksheets/sheet1.xml" and "worksheets/sheet1.xml"
            tgt = target[rid].lstrip("/")
            part = tgt if tgt.startswith("xl/") else "xl/" + tgt
            xml = z.read(part).decode("utf-8", "replace")
            row1 = re.search(r"<row[^>]*\sr=\"1\"[^>]*>(.*?)</row>", xml, re.S)
            by_col = {}
            if row1:
                for ref, attrs, body in re.findall(
                        r'<c r="([A-Z]+\d+)"([^>]*)>(.*?)</c>', row1.group(1), re.S):
                    v = re.search(r"<v>(.*?)</v>", body, re.S)
                    t = re.search(r"<t[^>]*>(.*?)</t>", body, re.S)
                    if 't="s"' in attrs and v:
                        val = shared[int(v.group(1))] if int(v.group(1)) < len(shared) else ""
                    else:
                        val = t.group(1) if t else (v.group(1) if v else "")
                    val = (val.replace("&amp;", "&").replace("&lt;", "<")
                              .replace("&gt;", ">").replace("&quot;", '"')).strip()
                    if val:
                        by_col[col_of(ref)] = val
            width = max(by_col) if by_col else 0
            out[name] = [by_col.get(i, "") for i in range(1, width + 1)]
    return out


def sheet_col(ws, head):
    """Column of a heading in this sheet's own row 1, or None.

    Everything written into a staged row is placed through here rather than through
    the list the header was built from. The sheet is then the single authority on
    where a heading is, so a value cannot land under a different one."""
    if not head:
        return None
    want = str(head).strip().lower()
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if v is not None and str(v).strip().lower() == want:
            return c
    return None


def header_drift(ws, cols):
    """Headings in this sheet that are not in the column the master keeps them in.

    Worth checking every run, because a staged value is placed by looking its heading
    up in this sheet. A header row left over from an older version of this script
    would therefore put new values where that old row says, not where the master
    wants them, and a row copied across would land a column short."""
    bad = []
    for i, head in enumerate(cols, start=1):
        if not head:
            continue
        at = sheet_col(ws, head)
        if at != i:
            bad.append((head, at, i))
    return bad


def master_headers():
    """Row 1 of each master sheet, so the staging file mirrors it exactly."""
    if not MASTER.exists():
        sys.exit("Cannot find the master workbook: %s" % MASTER)
    heads = header_row_from_xlsx(MASTER, SHEETS)
    heads["Posts"] = list(POST_COLS)
    return heads


def new_book(heads):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for name in SHEETS + ("Posts",):
        write_header(wb.create_sheet(name), heads[name])
    return wb


def write_header(ws, cols):
    """Master columns, a blank gutter, then the review columns."""
    grey = PatternFill("solid", fgColor="EFF4F5")
    green = PatternFill("solid", fgColor="E4F0EE")
    for i, h in enumerate(cols, start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = Font(bold=True)
        c.fill = grey
    start = len(cols) + 2                        # +2 leaves one empty gutter column
    for j, h in enumerate(REVIEW):
        c = ws.cell(row=1, column=start + j, value=h)
        c.font = Font(bold=True)
        c.fill = green
        c.alignment = Alignment(wrap_text=True)
        ws.column_dimensions[get_column_letter(start + j)].width = 18
    ws.freeze_panes = "A2"


def staged_ids(wb):
    """Every Netlify id already in the file, so nothing is staged twice."""
    seen = set()
    for ws in wb.worksheets:
        col = sheet_col(ws, ID_COL)
        if not col:
            continue
        for r in range(2, ws.max_row + 1):
            v = ws.cell(row=r, column=col).value
            if v:
                seen.add(str(v))
    return seen


def target_sheet(form_name, data):
    """Which sheet a submission belongs on."""
    if form_name == "cbeds-charter":
        return "Agent"                       # a signatory is listed on CBEDSync
    kind = (data.get("type") or "").strip().lower()
    return {"agent": "Agent", "project": "Project",
            "output": "Output", "post": "Posts"}.get(kind, "Posts")


# Fields were renamed so that Netlify's emails and CSV exports read as labels rather
# than as shorthand - "Organisation" not "Org", "Commitments" not "Commit",
# "Submitted-by" not "By". Submissions taken before each rename still carry the old
# name, so both spellings are accepted and nothing already collected is lost.
WAS_CALLED = {"organisation": "org", "commitments": "commit", "submitted-by": "by"}


def field(data, name):
    v = data.get(name)
    if v in (None, "") and name in WAS_CALLED:
        v = data.get(WAS_CALLED[name])
    return v


def entity_name(form_name, data):
    if form_name == "cbeds-charter":
        return (field(data, "organisation") or "").strip()
    return (data.get("name") or "").strip()


def stage(ws, heads, sub):
    """Append one submission as a row: master columns first, then review columns."""
    data = sub.get("data") or {}
    form = sub.get("form_name") or ""
    row = ws.max_row + 1

    def put_head(head, value):
        if not value:
            return
        c = sheet_col(ws, head)
        if c is not None:
            ws.cell(row=row, column=c, value=value)

    # the entity's name is always the sheet's first column
    ws.cell(row=row, column=1, value=entity_name(form, data))
    for src, head in FIELD_TO_HEAD.items():
        put_head(head, (data.get(src) or "").strip())
    # The chips arrive already split by what they are, and already carrying the exact
    # names the workbook uses, so each goes straight to its own columns. Whatever will
    # not fit goes to a review column rather than being dropped on the floor.
    def listed(name):
        return [x.strip() for x in (data.get(name) or "").split(",") if x.strip()]

    overflow = []
    for values, cols in ((listed("technologies"), TECH_HEADS),
                         (listed("related"), LINK_HEADS)):
        for head, value in zip(cols, values):
            put_head(head, value)
        overflow += values[len(cols):]
    for theme in listed("themes"):
        if sheet_col(ws, theme) is None:
            overflow.append(theme)          # no column for it on this sheet
        else:
            put_head(theme, THEME_MARK)
    # submissions taken before the chips carried canonical names
    overflow += listed("aspects")
    put_head(SOURCE_HEAD, "Public")          # the whole point: it came from outside

    commits = field(data, "commitments")
    if isinstance(commits, list):
        commits = ", ".join(commits)

    values = {
        "Review: Submitted": (sub.get("created_at") or "")[:19].replace("T", " "),
        "Review: Submitted by": field(data, "submitted-by") or field(data, "organisation") or "",
        "Review: Email": data.get("email") or "",
        "Review: Named lead": data.get("lead") or "",
        "Review: Commitments": commits or "",
        "Review: Unplaced": ", ".join(overflow),
        "Review: Form": form,
        ID_COL: str(sub.get("id") or ""),
    }
    for head, value in values.items():
        col = sheet_col(ws, head)
        if col and value:
            ws.cell(row=row, column=col, value=value)


def main():
    token = env("NETLIFY_TOKEN")
    if not token:
        sys.exit(
            "No NETLIFY_TOKEN found.\n\n"
            "1. Create a personal access token at\n"
            "   https://app.netlify.com/user/applications  ->  New access token\n"
            "2. Put it in the .env file next to this script:\n"
            "       NETLIFY_TOKEN=your-token-here\n"
            "   .env is git-ignored, so it stays on this computer.")
    site = env("NETLIFY_SITE", "cbeds.netlify.app")

    heads = master_headers()
    if STAGING.exists():
        wb = openpyxl.load_workbook(STAGING)
        drift = []
        for name in SHEETS:
            for head, at, want in header_drift(wb[name], heads[name]):
                drift.append("%s: %s is in column %s, the master has it in %s"
                             % (name, head,
                                get_column_letter(at) if at else "(missing)",
                                get_column_letter(want)))
        if drift:
            sys.exit(
                "%s does not match the master's columns:\n  %s\n\n"
                "Delete draft/%s and run this again to rebuild it. Every submission\n"
                "is fetched fresh from Netlify, so nothing is lost except anything\n"
                "written by hand in the Review columns - copy that out first if you\n"
                "want to keep it."
                % (STAGING.name, "\n  ".join(drift[:6]), STAGING.name))
    else:
        wb = new_book(heads)
    already = staged_ids(wb)

    site_id, site_name = resolve_site(site, token)
    print("  project: %s" % site_name)

    forms = api("/sites/%s/forms" % site_id, token)
    wanted = {f["name"]: f["id"] for f in forms
              if f.get("name") in ("cbeds-charter", "cbedsense-submission")}
    if not wanted:
        seen = ", ".join(sorted(str(f.get("name")) for f in forms)) or "(none)"
        sys.exit("No CBEDS forms on %s.\n"
                 "Forms this project has: %s\n"
                 "If that is empty, no deploy carrying the forms has gone out yet."
                 % (site_name, seen))
    print("  forms  : %s" % ", ".join(sorted(wanted)))

    added = skipped = 0
    for name, fid in sorted(wanted.items()):
        for sub in api("/forms/%s/submissions?per_page=100" % fid, token):
            if str(sub.get("id")) in already:
                skipped += 1
                continue
            sheet = target_sheet(name, sub.get("data") or {})
            stage(wb[sheet], heads[sheet], sub)
            added += 1

    wb.save(STAGING)
    print("OK  wrote %s" % STAGING.name)
    print("    new=%d  already staged=%d" % (added, skipped))
    if added:
        print("    Check them, then copy the master columns of the rows you approve")
        print("    into CBEDSync.xlsx. Source already says Public.")


if __name__ == "__main__":
    main()
